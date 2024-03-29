from fastapi import FastAPI, HTTPException

import pymorphy2
from utils import db
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

morph = pymorphy2.MorphAnalyzer()

app = FastAPI()


@app.get('/report')
async def create_report(user_id: int, category_id: int):
    week = await db.get_search_by_report_type('week')
    month = await db.get_search_by_report_type('month')
    month3 = await db.get_search_by_report_type('3month')
    priorities = await db.get_priorities_by_category_id(category_id)

    word_forms_data = {}
    for record in priorities:
        for word in record["query"].split(" "):
            normal_form = morph.parse(word)[0].normal_form
            try:

                word_forms_data[normal_form]["entries_count"] += 1
                word_forms_data[normal_form]["frequency"] = max(record["frequency"],
                                                                word_forms_data[normal_form]["frequency"])
                if word not in word_forms_data[normal_form]["word_forms"]:
                    word_forms_data[normal_form]["word_forms"].append(word)
            except KeyError:
                word_forms_data[normal_form] = {"name": normal_form, "word_forms": [word], "entries_count": 1,
                                                "frequency": record["frequency"]}
            except Exception as e:
                continue

    res = []
    for value in word_forms_data.values():
        value["word_forms"] = ", ".join(value["word_forms"])
        res.append(value)
    word_forms_data = res

    wb = Workbook(write_only=True)

    priorities_ws = wb.create_sheet("Запросы")
    priorities_ws.append(["Поисковый запрос", "Частотность", "Кол-во товаров", "Приоритет"])
    for record in priorities:
        record = dict(record)
        record.pop("search_id")
        priorities_ws.append(list(record.values()))

    word_forms_ws = wb.create_sheet("Словоформы")
    word_forms_ws.append(["Слово", "Словоформа", "Кол-во вхождений", "Суммарная частотность"])
    for record in word_forms_data:
        word_forms_ws.append(list(record.values()))

    filename = f"{user_id}-{category_id}.xlsx"
    wb.save(filename)
    return {"filename": filename}


@app.get("/seasonality")
async def get_seasonality(query: str):
    search_id = await db.get_search_id_by_query(query)
    if search_id is None:
        raise HTTPException(400)
    response_data = []
    report_types = ["week", "month", "3month"]
    for report_type in report_types:
        frequencies = await db.get_frequency_by_search_id_and_report_type(search_id, report_type)
        if len(frequencies) == 0:
            frequency_diff = 0
            frequency = 0
        elif len(frequencies) == 1:
            frequency_diff = 0
            frequency = frequencies[0]["frequency"]
        else:
            frequency_diff = (frequencies[1]["frequency"] - frequencies[0]["frequency"]) / frequencies[0][
                "frequency"] * 100
            frequency_diff = round(frequency_diff, 1)
            frequency = frequencies[1]["frequency"]
        response_data.append(
            {"date_type": report_type, "frequency": frequency, "diff": frequency_diff})
    return response_data
